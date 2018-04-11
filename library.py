from tkinter import *
from tkinter import ttk
import openpyxl as px, os

os.chdir('/Users/Yusuke/Documents/Python/materials')

wb = px.load_workbook('library.xlsx')
ws = wb.get_sheet_by_name('Library')

list2 = []
list3 = []
list4 = []

for col in range(1, 3716):
    list2.append(ws.cell(col, 3).value)
    list3.append(ws.cell(col, 15).value)

def button_click1():
    lb.delete(0, END)
    value1 = EditBox1.get()
    for i in range(0, 3715):
        index = list2[i].find(value1)
        if index != -1:
            if list3[i] == "貸出中":
                lb.insert(END, "貸出中：" + list2[i])
            else:
                lb.insert(END, list2[i])

def listbox_selected(event):
    show_selection()

def show_selection():
    for i in lb.curselection():
        value2 = lb.get(i)
        list4.append(value2)

def button_click2():
    value2 = EditBox2.get()
    for i in range(0, 3715):
        if list2[i] == list4[-1]:
            ws.cell(row=i+1, column=15).value = "貸出中"
            ws.cell(row=i+1, column=16).value = value2
            wb.save("library.xlsx")

def button_click3():
    for i in range(0, 3715):
        if "貸出中：" + list2[i] == list4[-1]:
            ws.cell(row=i+1, column=15).value = ""
            ws.cell(row=i+1, column=16).value = ""
            wb.save("library.xlsx")

if __name__ == '__main__':
    root = Tk()
    root.title('谷口ラボ')

    frame1 = ttk.Frame(root, padding=10)
    frame1.grid()

    frame2 = ttk.Frame(root, padding=10)
    frame2.grid()

    button1 = ttk.Button(frame2, text='検索', command=button_click1)
    button1.grid(row=0, column=1)

    button2 = ttk.Button(frame2, text='貸出', command=button_click2)
    button2.grid(row=0, column=4)

    button3 = ttk.Button(frame2, text='返却', command=button_click3)
    button3.grid(row=0, column=5)

    lb = Listbox(frame1, width =42, height=10)
    lb.insert(END)
    lb.bind('<<ListboxSelect>>', listbox_selected)
    lb.grid(row=0, column=0)

    EditBox1 = ttk.Entry(frame2)
    EditBox1.insert(END, "")
    EditBox1.grid(row=1, column=0, columnspan=3)

    EditBox2 = ttk.Entry(frame2)
    EditBox2.insert(END, "")
    EditBox2.grid(row=1, column=3, columnspan=3)

    root.mainloop()